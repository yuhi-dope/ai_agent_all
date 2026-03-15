"""ドキュメント生成エンジン（Excel/PDF/Word）"""
import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class ExcelGenerator:
    """Excel生成エンジン"""

    @staticmethod
    def generate_table(
        title: str,
        headers: list[str],
        rows: list[list[Any]],
        column_widths: list[int] | None = None,
    ) -> bytes:
        """テーブル形式のExcelを生成"""
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excelシート名は31文字まで

        # タイトル行
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center")

        # ヘッダー行
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # データ行
        for row_idx, row_data in enumerate(rows, 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # 列幅
        if column_widths:
            for col_idx, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col_idx)].width = width

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def generate_from_template(
        template_data: dict[str, Any],
    ) -> bytes:
        """テンプレートデータからExcelを生成（内訳書等）"""
        wb = Workbook()
        ws = wb.active

        title = template_data.get("title", "")
        ws.title = title[:31]

        # メタ情報
        meta = template_data.get("meta", {})
        row = 1
        for key, value in meta.items():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=str(value))
            row += 1

        row += 1  # 空行

        # テーブルデータ
        headers = template_data.get("headers", [])
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True)
        row += 1

        for item in template_data.get("rows", []):
            for col_idx, value in enumerate(item, 1):
                ws.cell(row=row, column=col_idx, value=value)
            row += 1

        # 合計行
        totals = template_data.get("totals", {})
        if totals:
            row += 1
            for key, value in totals.items():
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=value)
                ws.cell(row=row, column=1).font = Font(bold=True)
                row += 1

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
